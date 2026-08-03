"""
make_excel_dashboard.py
Builds an Excel dashboard (.xlsx) from scheme_performance data:
raw data sheet, KPIs, summary by fund house, and a bar chart.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

df = pd.read_csv("data/raw/07_scheme_performance.csv")

wb = Workbook()

# --- Sheet 1: raw data ---
ws = wb.active
ws.title = "Data"
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

# --- Sheet 2: summary by fund house (pivot-style) ---
summary = df.groupby("fund_house")["return_3yr_pct"].mean().round(2).reset_index()
summary.columns = ["Fund House", "Avg 3Yr Return %"]
ws2 = wb.create_sheet("Summary")
for r in dataframe_to_rows(summary, index=False, header=True):
    ws2.append(r)

# bar chart of the summary
chart = BarChart()
chart.title = "Avg 3-Year Return by Fund House"
chart.y_axis.title = "Return %"
data_ref = Reference(ws2, min_col=2, min_row=1, max_row=len(summary)+1)
cats_ref = Reference(ws2, min_col=1, min_row=2, max_row=len(summary)+1)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
ws2.add_chart(chart, "D2")

# --- Sheet 3: KPIs ---
ws3 = wb.create_sheet("Dashboard")
ws3["A1"] = "MUTUAL FUND PERFORMANCE DASHBOARD"
ws3["A3"] = "Total Funds";        ws3["B3"] = len(df)
ws3["A4"] = "Avg 3Yr Return %";   ws3["B4"] = round(df["return_3yr_pct"].mean(), 2)
ws3["A5"] = "Best 3Yr Return %";  ws3["B5"] = round(df["return_3yr_pct"].max(), 2)
ws3["A6"] = "Avg Expense Ratio %"; ws3["B6"] = round(df["expense_ratio_pct"].mean(), 2)

wb.move_sheet("Dashboard", -2)  # put Dashboard first
wb.save("Fund_Dashboard.xlsx")
print("Created Fund_Dashboard.xlsx")
